# accounts/views.py
import random
import string
import secrets
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.contrib.sites.models import Site
from urllib.parse import quote
from .forms import UserCreateForm, CustomUserCreationForm, UserEditForm
from .models import User, DeletedUser, Profile, CustomerProfile
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.contrib.auth import get_user_model
from django.shortcuts import render
from django.shortcuts import redirect

User = get_user_model() # This is already in the provided code but is good practice to include.

# Utility Functions
def is_admin(user):
    """Check if user is admin or superuser"""
    return user.is_superuser or user.groups.filter(name="Admin").exists()

# Authentication Views
def custom_login(request):
    """Handle custom login with role-based authentication"""
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        role = request.POST.get("role")
        remember_me = request.POST.get("remember")  # Get the value of the 'remember' checkbox

        user = authenticate(request, username=username, password=password)
        if user is not None:
            # Role-based validation
            if role == "admin" and user.is_superuser:
                login(request, user)
                if not remember_me:
                    request.session.set_expiry(0)  # Make session expire on browser close
                return redirect("accounts:dashboard_home")
            elif role == "staff" and user.groups.filter(name="Staff").exists():
                login(request, user)
                if not remember_me:
                    request.session.set_expiry(0) # Make session expire on browser close
                return redirect("accounts:dashboard_home")
            elif role == "customer" and user.groups.filter(name="Customer").exists():
                login(request, user)
                if not remember_me:
                    request.session.set_expiry(0) # Make session expire on browser close
                return redirect("accounts:dashboard_home")

            messages.error(request, "This account doesn't have the selected role")
        else:
            messages.error(request, "Invalid Username or Password")

        return redirect("accounts:login")

    return render(request, "accounts/login.html")

def custom_logout(request):
    """Handle user logout"""
    logout(request)
    return redirect("accounts:login")

@login_required
def dashboard_home(request):
    """Dashboard homepage"""
    user_first_name = request.user.first_name
    context = {
        "first_name": user_first_name,
    }
    return render(request, "accounts/dashboard_home.html", context)

def csrf_failure(request, reason=""):
    # Optional: Add a message to display on the login page
    messages.error(request, "Your session has expired. Please try logging in again.")
    # Redirect the user back to the login page
    return redirect('accounts:login')

# User Management Views
@login_required
def all_users(request):
    """Display list of all users"""
    # Use select_related to efficiently fetch related profiles
    users = User.objects.filter(is_deleted=False).select_related('account_profile', 'customer_profile')
    
    context = {
        "all_users": users,
        "first_name": request.user.first_name, 
    }
    
    return render(request, "accounts/all_users.html", context)

@login_required
@user_passes_test(is_admin)
def add_user(request):
    """Handle new user creation"""
    if request.method == "POST":
        form = UserCreateForm(request.POST)
        
        if not form.is_valid():
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
            
            context = {
                "form": form,
                "first_name": request.user.first_name, 
            }
            return render(request, "accounts/add_user.html", context)
        
        try:
            user, raw_password, login_link = form.save()
            
            messages.success(request, f"User {user.username} created successfully! Password: {raw_password}. Login at: {login_link}")
            return redirect("accounts:all_users")
            
        except Exception as e:
            messages.error(request, f"Error creating user: {str(e)}")
            
    else:
        form = UserCreateForm()
    
    context = {
        "form": form,
        "first_name": request.user.first_name,
    }

    return render(request, "accounts/add_user.html", context)

@login_required
@user_passes_test(is_admin)
def send_whatsapp_welcome(request, user_id):
    """Send welcome message via WhatsApp"""
    user = get_object_or_404(User, id=user_id)

    # Get user profile data using the same logic as the template filter
    whatsapp_number = None
    raw_password = None
    if hasattr(user, 'account_profile') and user.account_profile:
        whatsapp_number = user.account_profile.whatsapp_number
        raw_password = user.account_profile.raw_password
    elif hasattr(user, 'customer_profile') and user.customer_profile:
        whatsapp_number = user.customer_profile.whatsapp_number
        raw_password = user.customer_profile.raw_password

    if not whatsapp_number:
        messages.error(request, "User doesn't have a WhatsApp number")
        return redirect("accounts:all_users")

    current_site = Site.objects.get_current()
    login_link = f"http://{current_site.domain}/accounts/login/"

    message = f"Hello {user.first_name or user.username}, Welcome to Akshardeep Offset Printers ERP\n\n"
    message += f"Here is Your Login ID: {user.username}\n"
    message += f"Password: {raw_password or 'Please contact admin for password'}\n"
    message += f"Click on this link to Login: {login_link}"

    whatsapp_number = str(whatsapp_number).strip()
    whatsapp_number = ''.join(c for c in whatsapp_number if c.isdigit() or c == '+')

    if not whatsapp_number.startswith('+'):
        whatsapp_number = '+91' + whatsapp_number

    whatsapp_url = f"https://wa.me/{whatsapp_number}?text={quote(message)}"

    return redirect(whatsapp_url)

@login_required
@user_passes_test(is_admin)
def edit_user(request, user_id):
    """Edit existing user details"""
    user = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save() # The form's save method now handles the password change.
            messages.success(request, "User updated successfully")
            return redirect("accounts:all_users")
    else:
        form = UserEditForm(instance=user)

    context = {
        "form": form,
        "user": user,
        "first_name": request.user.first_name,
    }

    return render(request, "accounts/edit_user.html", context)

# Delete/Restore Views
@login_required
@user_passes_test(is_admin)
def delete_user(request, user_id):
    """Soft delete a user (move to recycle bin)"""
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        role = "Admin" if user.is_superuser else \
            user.groups.first().name if user.groups.exists() else "Customer"

        whatsapp_number = None
        if hasattr(user, 'account_profile'):
            whatsapp_number = user.account_profile.whatsapp_number
        elif hasattr(user, 'customer_profile'):
            whatsapp_number = user.customer_profile.whatsapp_number

        DeletedUser.objects.create(
            original_id=user.id, # Save the original ID for restoration
            first_name=user.first_name,
            last_name=user.last_name,
            username=user.username,
            email=user.email,
            whatsapp_number=whatsapp_number,
            role=role
        )

        user.is_deleted = True
        user.deleted_at = timezone.now()
        user.save()

        messages.success(request, f"User {user.username} moved to recycle bin")
        return redirect('accounts:all_users')

    return render(request, "accounts/confirm_delete.html", {"user": user})

@login_required
@user_passes_test(is_admin)
def permanent_delete_user(request, deleted_user_id):
    """Permanently delete a user from the recycle bin."""
    try:
        deleted_user = get_object_or_404(DeletedUser, id=deleted_user_id)
        
        if request.method == 'POST':
            username = deleted_user.username
            original_user_id = deleted_user.original_id
            
            # Find and hard delete the original user if it exists
            if original_user_id:
                try:
                    user_to_delete = User.objects.get(id=original_user_id)
                    user_to_delete.hard_delete()
                    messages.success(request, f"User {username} and associated records permanently deleted.")
                except User.DoesNotExist:
                    messages.success(request, f"Original user record for {username} not found. Deleting from recycle bin.")
            
            deleted_user.delete() # Always delete the record from DeletedUser model

            return redirect('accounts:recycle_bin')

        return render(request, "accounts/confirm_permanent_delete.html", {"user": deleted_user})

    except Exception as e:
        messages.error(request, f"Error deleting user: {str(e)}")
        return redirect('accounts:recycle_bin')

@login_required
@user_passes_test(is_admin)
def restore_user(request, deleted_id):
    """Restore a user from recycle bin"""
    deleted_user = get_object_or_404(DeletedUser, id=deleted_id)

    if request.method == 'POST':
        # Find the original user record
        try:
            user = User.objects.get(id=deleted_user.original_id)
            user.is_deleted = False
            user.deleted_at = None
            user.save()
            messages.success(request, f"User {user.username} restored successfully")

        except User.DoesNotExist:
            messages.error(request, f"Original user record not found for {deleted_user.username}. Cannot restore.")

        deleted_user.delete()
        return redirect('accounts:all_users')

    return render(request, "accounts/confirm_restore.html", {"user": deleted_user})

@login_required
def recycle_bin(request):
    """Show deleted users"""
    deleted_users = DeletedUser.objects.all().order_by('-deleted_at')
    return render(request, "accounts/recycle_bin.html", {"deleted_users": deleted_users})

# Import/Export Views
@login_required
@user_passes_test(is_admin)
def download_excel(request):
    """Export users to Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    headers = ["First Name", "Last Name", "Username", "Email", "Password", "Role", "WhatsApp", "Press Name", "Status"]
    ws.append(headers)

    users = User.objects.all().select_related('account_profile', 'customer_profile')
    for user in users:
        role = "Admin" if user.is_superuser else \
            user.groups.first().name if user.groups.exists() else "User"

        whatsapp = None
        password = None
        press_name = None

        if hasattr(user, 'account_profile'):
            whatsapp = user.account_profile.whatsapp_number
            password = user.account_profile.raw_password
            press_name = user.account_profile.press_name
        elif hasattr(user, 'customer_profile'):
            whatsapp = user.customer_profile.whatsapp_number
            password = user.customer_profile.raw_password
            press_name = user.customer_profile.press_name


        whatsapp = whatsapp or "-"
        password = password or "-"
        press_name = press_name or "-"

        status = "Active" if not user.is_deleted else "Deleted"

        ws.append([
            user.first_name,
            user.last_name,
            user.username,
            user.email,
            password,
            role,
            whatsapp,
            press_name,
            status
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="users.xlsx"'
    wb.save(response)
    return response

@login_required
@user_passes_test(is_admin)
def download_pdf(request):
    """Export users to PDF"""
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="users.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph("User List", styles["Title"]))

    data = [["First Name", "Last Name", "Username", "Email", "Password", "Role", "Status"]]

    users = User.objects.all().select_related('account_profile', 'customer_profile')
    for user in users:
        role = "Admin" if user.is_superuser else \
            user.groups.first().name if user.groups.exists() else "User"

        password = None
        if hasattr(user, 'account_profile'):
            password = user.account_profile.raw_password
        elif hasattr(user, 'customer_profile'):
            password = user.customer_profile.raw_password

        password = password or "-"

        status = "Active" if not user.is_deleted else "Deleted"

        data.append([
            user.first_name,
            user.last_name,
            user.username,
            user.email,
            password,
            role,
            status
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.grey),
        ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0), (-1,-0), 12),
        ("BACKGROUND", (0,1), (-1,-1), colors.beige),
        ("GRID", (0,0), (-1,-1), 1, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)
    return response

@login_required
@user_passes_test(is_admin)
def upload_users(request):
    """Handle user uploads via Excel file"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            created_users = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 4:
                    first_name, last_name, email, whatsapp = row[:4]
                    if not email:
                        continue

                    username = email.split("@")[0]
                    password = generate_password()

                    if not User.objects.filter(username=username).exists():
                        user = User.objects.create_user(
                            username=username,
                            email=email,
                            first_name=first_name,
                            last_name=last_name,
                            password=password
                        )

                        # Check for and update the correct profile type based on the
                        # signal logic. Since the signal creates the profile,
                        # we just need to update it.
                        if hasattr(user, 'account_profile'):
                            user.account_profile.raw_password = password
                            user.account_profile.whatsapp_number = whatsapp
                            user.account_profile.save()
                        elif hasattr(user, 'customer_profile'):
                            user.customer_profile.raw_password = password
                            user.customer_profile.whatsapp_number = whatsapp
                            user.customer_profile.save()

                        created_users += 1

            return JsonResponse({"success": True, "message": f"Successfully created {created_users} users"})

        except Exception as e:
            return JsonResponse({"success": False, "message": f"Error processing file: {str(e)}"})

    return JsonResponse({"success": False, "message": "No file provided"})